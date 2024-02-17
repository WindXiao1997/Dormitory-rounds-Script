# 导入所需的库
import pandas as pd
import openpyxl

# 读取Excel文件中的数据表
df = pd.read_excel("通行记录20231101082425.xlsx", sheet_name="数据表")

# 定义四种异常情况的列表
night_out = [] # 夜不归宿
late_back = [] # 晚归
Morning_abnormal = [] # 早上异常出入
Noon_abnormal = [] # 中午异常出入

# 遍历数据表中的每一行
for i, row in df.iterrows():
    # 获取姓名、出入口、通行时间等信息
    name = row["姓名"]
    inout = row["出入类型"]
    time = row["通行时间"]

    # 导入datetime模块
    import datetime

    # 把time变量转换成datetime对象
    time = datetime.datetime.strptime(time, "%Y/%m/%d %H:%M:%S")

    # 判断是否是第一条记录或者最后一条记录
    first = i == 0 or name != df.loc[i-1, "姓名"]
    last = i == len(df) - 1 or name != df.loc[i+1, "姓名"]

    # 判断是否是夜不归宿
    if last and inout == "出口":
        night_out.append(name)
        df.loc[i, "异常"] = "夜不归宿" # 在数据表中添加一列异常，标记异常类型

    # 判断是否是晚归
    if last and inout == "入口" and time.hour >= 23.5:
        late_back.append(name)
        df.loc[i, "异常"] = "晚归"

    # 判断是否是早上异常出入
    if first and inout == "入口":
        Morning_abnormal.append(name)
        df.loc[i, "异常"] = "早上异常出入"

    # 判断是否是中午异常出入
    if last and time.hour < 12:
        Noon_abnormal.append(name)
        df.loc[i, "异常"] = "中午异常出入"

# 创建一个新的工作簿对象，并命名为Result.xlsx
wb = openpyxl.Workbook()
wb.save("Result.xlsx")

# 打开Result.xlsx文件，获取当前的工作表对象，并删除它（因为我们不需要它）
wb = openpyxl.load_workbook("Result.xlsx")
ws = wb.active
wb.remove(ws)

# 创建四个新的Sheet对象，分别命名为夜不归宿、晚归、早上异常出入、中午异常出入，并复制数据表的表头到每个Sheet中
ws1 = wb.create_sheet("夜不归宿")
ws2 = wb.create_sheet("晚归")
ws3 = wb.create_sheet("早上异常出入")
ws4 = wb.create_sheet("中午异常出入")
for j in range(1, 27):
    ws1.cell(1, j).value = df.columns[j-1]
    ws2.cell(1, j).value = df.columns[j-1]
    ws3.cell(1, j).value = df.columns[j-1]
    ws4.cell(1, j).value = df.columns[j-1]

# 定义四个计数器，分别记录每个Sheet中的行数
row1 = 2
row2 = 2
row3 = 2
row4 = 2

# 遍历数据表中的每一行，根据姓名和异常类型，将对应的记录复制到相应的Sheet中，并更新计数器
for i in range(len(df)):
    # 获取姓名和异常类型等信息
    name = df.loc[i, "姓名"]
    exception = df.loc[i, "异常"]

    # 判断是否是夜不归宿，如果是，则复制到夜不归宿Sheet中，并更新row1计数器
    if name in night_out:
        for j in range(1, 28):
            ws1.cell(row1, j).value = df.iloc[i, j-1]
        row1 += 1

    # 判断是否是晚归，如果是，则复制到晚归Sheet中，并更新row2计数器
    if name in late_back:
        for j in range(1, 28):
            ws2.cell(row2, j).value = df.iloc[i, j-1]
        row2 += 1

    # 判断是否是早上异常出入，如果是，则复制到早上异常出入Sheet中，并更新row3计数器
    if name in Morning_abnormal:
        for j in range(1, 28):
            ws3.cell(row3, j).value = df.iloc[i, j-1]
        row3 += 1

    # 判断是否是中午异常出入，如果是，则复制到中午异常出入Sheet中，并更新row4计数器
    if name in Noon_abnormal:
        for j in range(1, 28):
            ws4.cell(row4, j).value = df.iloc[i, j-1]
        row4 += 1

# 保存Result.xlsx文件
wb.save("Result.xlsx")


# 打印四种情况的名单
print("夜不归宿的人员有：", ", ".join(night_out))
print("晚归的人员有：", ", ".join(late_back))
print("早上第一条记录是进入记录的异常人员有：", ", ".join(Morning_abnormal))
print("中午以后没有离开宿舍的异常人员有：", ", ".join(Noon_abnormal))
