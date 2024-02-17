# 导入pandas和openpyxl库
import pandas as pd
import openpyxl
import os

# 定义目录路径，使用当前工作目录和相对路径
directory = os.getcwd()

# 读取list.xlsx文件中的数据表
list_df = pd.read_excel(os.path.join(directory, "list.xlsx"), sheet_name="Sheet1")

# 定义一个字典，用来存储宿舍号和名单的对应关系
dorm_dict = {}

# 遍历数据表中的每一行
for i, row in list_df.iterrows():
    # 获取宿舍号
    dorm = row.iloc[0]
    # 获取名单，去掉空值和重复值
    names = row[1:].dropna().unique()
    # 遍历名单中的每一个名字
    for name in names:
        # 把名字和宿舍号的对应关系存入字典中
        dorm_dict[name] = dorm

# 读取通行记录.xlsx文件中的数据表
record_df = pd.read_excel(os.path.join(directory, "通行记录20231102135920.xlsx"), sheet_name="数据表")

# 定义四种异常情况的列表
night_out = [] # 夜不归宿
late_back = [] # 晚归
Morning_abnormal = [] # 早上异常出入
Noon_abnormal = [] # 中午异常出入

# 遍历数据表中的每一行
for i, row in record_df.iterrows():
    # 获取姓名、出入口、通行时间等信息
    name = row["姓名"]
    inout = row["出入类型"]
    time = row["通行时间"]

    # 导入datetime模块
    import datetime

    # 把time变量转换成datetime对象
    time = datetime.datetime.strptime(time, "%Y/%m/%d %H:%M:%S")

    # 判断是否是第一条记录或者最后一条记录
    first = i == 0 or name != record_df.loc[i-1, "姓名"]
    last = i == len(record_df) - 1 or name != record_df.loc[i+1, "姓名"]

    # 判断是否是夜不归宿
    if last and inout == "出口":
        night_out.append(name)
        record_df.loc[i, "异常"] = "夜不归宿" # 在数据表中添加一列异常，标记异常类型

    # 判断是否是晚归
    if last and inout == "入口" and time.hour >= 23:
        late_back.append(name)
        record_df.loc[i, "异常"] = "晚归"

    # 判断是否是早上异常出入
    if first and inout == "入口":
        Morning_abnormal.append(name)
        record_df.loc[i, "异常"] = "早上异常出入"

    # 判断是否是中午异常出入
    if last and time.hour < 12:
        Noon_abnormal.append(name)
        record_df.loc[i, "异常"] = "中午异常出入"

# 创建一个新的工作簿对象，并命名为通行记录+Result.xlsx
wb = openpyxl.Workbook()
wb.save(os.path.join(directory, "通行记录Result.xlsx"))

# 打开通行记录+Result.xlsx文件，获取当前的工作表对象，并删除它（因为我们不需要它）
wb = openpyxl.load_workbook(os.path.join(directory, "通行记录Result.xlsx"))
ws = wb.active
wb.remove(ws)

# 创建四个新的Sheet对象，分别命名为夜不归宿、晚归、早上异常出入、中午异常出入，并复制数据表的表头到每个Sheet中
ws1 = wb.create_sheet("夜不归宿")
ws2 = wb.create_sheet("晚归")
ws3 = wb.create_sheet("早上异常出入")
ws4 = wb.create_sheet("中午异常出入")
for j in range(1, 28):
    ws1.cell(1, j).value = record_df.columns[j-1]
    ws2.cell(1, j).value = record_df.columns[j-1]
    ws3.cell(1, j).value = record_df.columns[j-1]
    ws4.cell(1, j).value = record_df.columns[j-1]

# 定义四个计数器，分别记录每个Sheet中的行数
row1 = 2
row2 = 2
row3 = 2
row4 = 2

# 遍历数据表中的每一行，根据姓名和异常类型，将对应的记录复制到相应的Sheet中，并更新计数器
for i in range(len(record_df)):
    # 获取姓名和异常类型等信息
    name = record_df.loc[i, "姓名"]
    exception = record_df.loc[i, "异常"]

    # 判断是否是夜不归宿，如果是，则复制到夜不归宿Sheet中，并更新row1计数器
    if name in night_out:
        for j in range(1, 28):
            ws1.cell(row1, j).value = record_df.iloc[i, j-1]
        row1 += 1

    # 判断是否是晚归，如果是，则复制到晚归Sheet中，并更新row2计数器
    if name in late_back:
        for j in range(1, 28):
            ws2.cell(row2, j).value = record_df.iloc[i, j-1]
        row2 += 1

    # 判断是否是早上异常出入，如果是，则复制到早上异常出入Sheet中，并更新row3计数器
    if name in Morning_abnormal:
        for j in range(1, 28):
            ws3.cell(row3, j).value = record_df.iloc[i, j-1]
        row3 += 1

    # 判断是否是中午异常出入，如果是，则复制到中午异常出入Sheet中，并更新row4计数器
    if name in Noon_abnormal:
        for j in range(1, 28):
            ws4.cell(row4, j).value = record_df.iloc[i, j-1]
        row4 += 1

    # 在表格中添加一列宿舍号，列头为"Dorm"
    for ws in [ws1, ws2, ws3, ws4]:
        ws.insert_cols(2)
        ws.cell(1, 2).value = "Dorm"
        # 遍历表格中的每一行
        for row in range(2, ws.max_row + 1):
            # 获取姓名
            name = ws.cell(row, 1).value
            # 根据字典中的对应关系，获取宿舍号
            dorm = dorm_dict.get(name, "Unknown")
            # 在表格中写入宿舍号
            ws.cell(row, 2).value = dorm

# 保存通行记录+Result.xlsx文件
wb.save(os.path.join(directory, "通行记录Result.xlsx"))

# 打印处理成功的信息和夜不归宿的人员名单
print(f"Processed 通行记录.xlsx and saved as 通行记录Result.xlsx")
print("夜不归宿的人员有：", ", ".join(night_out))
print("晚归的人员有：", ", ".join(late_back))
print("早上第一条记录是进入记录的异常人员有：", ", ".join(Morning_abnormal))
print("中午以后没有离开宿舍的异常人员有：", ", ".join(Noon_abnormal))
